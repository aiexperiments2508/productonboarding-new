"""The workbook: one sheet per branch, plus the two sheets that explain it.

This is the format a supplier will actually open, so it is the one that gets to
stop a mistake before it is made rather than report it afterwards. Three things
it does that a CSV cannot:

*   **The GTIN column is text.** Excel reads ``05012345600018`` as a number,
    drops the leading zero, and then offers to display the rest in scientific
    notation. The zero is significant. So the column carries an explicit ``@``
    format and every cell is written as a string.
*   **Enumerated columns get a dropdown**, listing the values this catalog has
    actually used. Not a hard constraint - a supplier introducing a genuinely
    new plug type is not making a mistake - so the validation warns rather than
    stops, and the values come from the catalog rather than from a list kept by
    hand.
*   **Required columns look required**, and safety-class ones look different
    again, because "a person reviews any change to this" is not the same
    statement as "this is required".

openpyxl, and nothing else. See ``requirements-datapack.txt`` for why this one
dependency is taken and the Word one is not.
"""

from __future__ import annotations

import io

from sc.datapack.sample import NOTE_COLUMN, NOT_REPRESENTABLE, Example
from sc.datapack.schema import LIST_SEPARATOR, Pack, Sheet

#: Where the dropdown source lists live. Hidden, because it is scaffolding.
LISTS_SHEET = "Lists"

#: Excel refuses a sheet name over 31 characters or containing []:*?/\\ .
MAX_SHEET_NAME = 31

#: Above this many options a dropdown is worse than a free cell.
MAX_DROPDOWN = 40


def _unavailable() -> str:
    return ("openpyxl is not installed, so the workbook was not written. "
            "pip install -r requirements-datapack.txt")


def available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def _safe_name(label: str, taken: set[str]) -> str:
    name = "".join(ch for ch in label if ch not in "[]:*?/\\")[:MAX_SHEET_NAME]
    candidate, n = name, 2
    while candidate in taken:
        candidate = f"{name[:MAX_SHEET_NAME - 2]}~{n}"
        n += 1
    taken.add(candidate)
    return candidate


def write(pack: Pack, examples: dict[str, Example] | None = None) -> bytes:
    """The whole pack as one workbook. Raises ImportError without openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    examples = examples or {}
    book = Workbook()
    book.remove(book.active)

    header = Font(bold=True, size=10)
    machine = Font(name="Consolas", size=9, color="666666")
    marker = Font(size=8, italic=True, color="666666")
    required_fill = PatternFill("solid", fgColor="FDF3D8")
    safety_fill = PatternFill("solid", fgColor="FADBD8")
    wrap = Alignment(wrap_text=True, vertical="top")

    readme = book.create_sheet("Read me first")
    _write_readme(readme, pack, examples, header, marker, wrap)

    taken: set[str] = {"Read me first"}
    lists = book.create_sheet(LISTS_SHEET)
    list_ranges = _write_lists(lists, pack, get_column_letter)
    lists.sheet_state = "hidden"

    categories = book.create_sheet("Categories")
    _write_categories(categories, pack, header)
    taken.add("Categories")
    taken.add(LISTS_SHEET)

    for sheet in pack.sheets:
        ws = book.create_sheet(_safe_name(sheet.label, taken))
        example = examples.get(sheet.branch)
        _write_sheet(ws, sheet, example, pack, list_ranges,
                     fonts=(header, machine, marker),
                     fills=(required_fill, safety_fill), wrap=wrap,
                     letter=get_column_letter, Validation=DataValidation)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _write_sheet(ws, sheet: Sheet, example: Example | None, pack: Pack,
                 list_ranges: dict[str, str], *, fonts, fills, wrap,
                 letter, Validation) -> None:
    header, machine, marker = fonts
    required_fill, safety_fill = fills
    with_note = example is not None and bool(example.rows)

    columns = list(sheet.columns)
    names = [c.name for c in columns]
    if with_note:
        names.append(NOTE_COLUMN)

    for index, column in enumerate(columns, start=1):
        col = letter(index)
        ws.cell(row=1, column=index, value=column.name).font = machine
        title = ws.cell(row=2, column=index, value=column.heading)
        title.font = header
        title.alignment = wrap
        note = ws.cell(row=3, column=index, value=column.marker())
        note.font = marker
        note.alignment = wrap

        if column.safety:
            title.fill = safety_fill
        elif column.required:
            title.fill = required_fill

        ws.column_dimensions[col].width = _width(column)

        # The GTIN trap, and every other string column that must not be
        # helpfully reinterpreted.
        if column.dtype == "str" or column.kind in ("identity", "image"):
            for row in range(4, 400):
                ws.cell(row=row, column=index).number_format = "@"

        values = pack.vocabulary.get(column.name)
        if values and column.dtype == "str" and len(values) <= MAX_DROPDOWN:
            ref = list_ranges.get(column.name)
            if ref:
                # allow_blank, and a warning rather than a stop: these are the
                # values the catalog has used, not the only lawful ones, and
                # refusing a new plug type would be the template overruling the
                # rules engine.
                validation = Validation(type="list", formula1=ref,
                                        allow_blank=True, showErrorMessage=False)
                ws.add_data_validation(validation)
                validation.add(f"{col}4:{col}400")

    if with_note:
        index = len(columns) + 1
        ws.cell(row=1, column=index, value=NOTE_COLUMN).font = machine
        ws.cell(row=2, column=index,
                value="What this row demonstrates").font = header
        ws.cell(row=3, column=index,
                value="ignored on the way in").font = marker
        ws.column_dimensions[letter(index)].width = 60

    for offset, row in enumerate(example.rows if example else [], start=4):
        for index, name in enumerate(names, start=1):
            ws.cell(row=offset, column=index, value=row.get(name, ""))

    ws.freeze_panes = "A4"


def _width(column) -> int:
    if column.kind == "identity":
        return 22
    if column.kind == "image":
        return 26
    if column.dtype.startswith("list["):
        return 38
    return max(14, min(30, len(column.heading) + 4))


def _write_lists(ws, pack: Pack, letter) -> dict[str, str]:
    """The hidden source lists, and where each one lives."""
    ranges: dict[str, str] = {}
    for index, (path, values) in enumerate(sorted(pack.vocabulary.items()),
                                           start=1):
        if len(values) > MAX_DROPDOWN:
            continue
        col = letter(index)
        ws.cell(row=1, column=index, value=path)
        for offset, value in enumerate(values, start=2):
            ws.cell(row=offset, column=index, value=value)
        ranges[path] = f"={LISTS_SHEET}!${col}$2:${col}${len(values) + 1}"
    return ranges


def _write_categories(ws, pack: Pack, header) -> None:
    ws.cell(row=1, column=1, value="Category code").font = header
    ws.cell(row=1, column=2, value="Where it sits").font = header
    ws.cell(row=1, column=3, value="Sheet").font = header
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 26
    row = 2
    for sheet in pack.sheets:
        for leaf in sheet.leaves:
            ws.cell(row=row, column=1, value=leaf)
            ws.cell(row=row, column=2, value=pack.taxonomy.get(leaf, leaf))
            ws.cell(row=row, column=3, value=sheet.label)
            row += 1
    ws.freeze_panes = "A2"


def _write_readme(ws, pack: Pack, examples: dict[str, Example],
                  header, marker, wrap) -> None:
    ws.column_dimensions["A"].width = 110
    lines: list[tuple[str, object]] = [
        (f"{pack.fascia} supplier product feed", header),
        ("", None),
        ("One sheet per part of the assortment. Fill in the sheet your line "
         "belongs to and send the workbook back through the vendor portal, "
         "inside a .zip with your photographs in an images/ folder.", wrap),
        ("", None),
        ("How to read a sheet", header),
        ("Row 1 is the machine name and is the only row we read. Row 2 is the "
         "label. Row 3 says whether the column is required, by which sales "
         "channel, and anything that will otherwise go wrong. Start typing on "
         "row 4.", wrap),
        ("", None),
        ("Four things a spreadsheet gets wrong on your behalf", header),
        ("1. GTIN is text. A leading zero is part of the number and Excel will "
         "remove it if the column is numeric. This column is already "
         "formatted as text; please keep it that way.", wrap),
        ("2. Units live in the header, never in the cell. Write 65, not 65 W.", wrap),
        ("3. Percentages are plain numbers. Write 90, not 90%.", wrap),
        (f"4. Lists go in one cell separated by '{LIST_SEPARATOR.strip()}'. "
         "Where the header says the order carries meaning - ingredients, "
         "fibre composition, INCI - that order is a legal declaration. Please "
         "do not sort it.", wrap),
        ("", None),
        ("What the shaded headings mean", header),
        ("A pale gold heading is required by at least one sales channel and "
         "the listing will not go live without it. A pale red heading is a "
         "safety-class attribute: a person reviews every change to one, and a "
         "value we have to infer rather than read blocks publication instead "
         "of degrading it.", wrap),
        ("", None),
        ("The worked examples", header),
        ("Every sheet is filled in with real lines from our catalogue, and the "
         "last few rows of each are deliberately wrong. The final column says "
         "what is wrong with each one and what we will say about it. Delete "
         "the example rows before you send the workbook back.", wrap),
    ]

    unshown: dict[str, str] = {}
    for example in examples.values():
        unshown.update(example.unshown)
    if unshown or NOT_REPRESENTABLE:
        lines.append(("", None))
        lines.append(("What the examples do not show", header))
        for defect, why in NOT_REPRESENTABLE.items():
            lines.append((f"{defect}: {why}.", wrap))
        for defect, why in sorted(unshown.items()):
            lines.append((f"{defect}: {why}.", wrap))

    for index, (text, style) in enumerate(lines, start=1):
        cell = ws.cell(row=index, column=1, value=text)
        if style is header:
            cell.font = header
        elif style is wrap:
            cell.alignment = wrap
