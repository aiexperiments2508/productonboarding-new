"""The supplier data pack, and the one thing it must never become.

A template is a statement about what a product record is, and the system
already has two: the attribute registry, and the checks that read it. A third
written by hand would drift within a release and drift silently - a supplier
would fill in a column nothing reads, or omit one that holds their launch.

So every assertion here is made against the *derivation* rather than against a
list of column names. A test that pinned the Grocery columns to a literal would
be exactly the third statement this file exists to prevent: it would pass while
disagreeing with the registry, and the first thing it would disagree about is
the attribute somebody added last week.

The trailing-dot case has its own test and is not a detail. ``applies_to``
holds taxonomy prefixes and five attributes are named leaf by leaf, so asking
for a branch's columns by branch key rather than by leaf silently returns only
the universal four - a template that asks a supplier for nothing.
"""

from __future__ import annotations

import io
import json
import os
import zipfile
from xml.etree import ElementTree

import pytest

os.environ.setdefault("DB_PATH", "data/test_datapack.db")

from sc import db  # noqa: E402
from sc.datapack import sample as sample_mod  # noqa: E402
from sc.datapack import schema  # noqa: E402
from sc.datapack.writers import csv_txt, jsonschema, specdoc, workbook  # noqa: E402
from sc.estate.defects import Defect  # noqa: E402
from sc.replay import tape  # noqa: E402
from sc.state import baseline as baseline_mod  # noqa: E402
from sc.state.baseline import applies_to_category  # noqa: E402

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


@pytest.fixture(autouse=True)
def fresh():
    db.init_db(drop=True)
    tape.load_tape(reset=True)
    yield
    db.close()


@pytest.fixture
def base():
    return baseline_mod.get()


@pytest.fixture
def pack(base):
    return schema.build(base)


# ---------------------------------------------------------------------------
# The columns are derived, not written down
# ---------------------------------------------------------------------------


def test_every_branch_gets_a_sheet_and_every_sheet_has_leaves(pack, base):
    assert pack.sheets, "a retailer that trades nothing has no pack"
    for sheet in pack.sheets:
        assert sheet.leaves, f"{sheet.branch} has no categories"
        assert all(leaf.count(".") == 2 for leaf in sheet.leaves)


def test_attribute_columns_are_exactly_what_the_registry_says_applies(pack, base):
    """Against the predicate, never against a list.

    An attribute belongs on a branch's sheet when it applies to at least one of
    that branch's categories - which is the same question
    ``applicable_attributes`` asks before it reports a gap. Two answers to it
    would mean a template that asks for a value no check reads, or a check that
    reports a value no template asked for.
    """
    for sheet in pack.sheets:
        drawn = {c.name for c in sheet.columns if c.kind == "attribute"}
        expected = {
            path for path, definition in base.attr_defs.items()
            if any(applies_to_category(definition, leaf) for leaf in sheet.leaves)
        }
        assert drawn == expected, sheet.branch


def test_a_saucepan_is_never_asked_for_a_wattage(pack, base):
    """The leaf-level case, pinned.

    ``specs.power_w`` is named leaf by leaf because a kettle is mains and a
    saucepan is not, and both are ``home.``. The column belongs on the Home
    sheet and must say which categories it covers, or a supplier fills in a
    wattage for a pan and four channels then require it.
    """
    home = pack.sheet("home")
    assert home is not None
    power = home.column("specs.power_w")
    assert power is not None, "the branch does have mains appliances in it"
    assert power.only_leaves, "a column that covers only some leaves must say so"
    assert len(power.only_leaves) < len(home.leaves)
    for leaf in home.leaves:
        covered = applies_to_category(base.attr_defs["specs.power_w"], leaf)
        assert covered == (leaf in power.only_leaves)


def test_asking_by_branch_key_rather_than_by_leaf_finds_almost_nothing(base):
    """The trailing-dot trap, pinned so nobody 'fixes' the prefix match.

    ``"food".startswith("food.")`` is false, so a derivation that passed the
    branch key straight to the predicate would quietly produce a template with
    only the attributes that apply everywhere. That template would look
    plausible and ask a grocery supplier for no ingredients.
    """
    universal = {p for p, d in base.attr_defs.items() if not d.applies_to}
    by_key = {p for p, d in base.attr_defs.items()
              if applies_to_category(d, "food")}
    assert by_key == universal
    # And the real derivation does not do that.
    food = schema.columns_for("food", base)
    drawn = {c.name for c in food if c.kind == "attribute"}
    assert "food.ingredients" in drawn
    assert drawn > universal


def test_image_columns_come_from_the_profile(pack, base):
    branches = schema.branches_of(base)
    for sheet in pack.sheets:
        required = {c.name.split(".", 1)[1] for c in sheet.columns
                    if c.kind == "image" and c.required}
        assert required == set(branches[sheet.branch]["required_media"])


def test_changing_the_profile_changes_the_pack(base, monkeypatch):
    """The 'nothing is hardcoded' claim, tested rather than asserted."""
    before = {c.name for c in schema.columns_for("food", base) if c.kind == "image"}
    profile = dict(base.catalog.profile)
    branches = {k: dict(v) for k, v in profile["branches"].items()}
    branches["food"] = {**branches["food"], "required_media": ["HERO"]}
    monkeypatch.setattr(base.catalog, "profile",
                        {**profile, "branches": branches})
    after = {c.name for c in schema.columns_for("food", base) if c.kind == "image"}
    assert before != after
    assert "image.HERO" in after


def test_required_columns_name_the_channels_that_require_them(pack):
    for sheet in pack.sheets:
        for column in sheet.columns:
            if column.kind != "attribute":
                continue
            assert bool(column.required_for) == column.required
            if column.required_for:
                assert "required by" in column.marker()


def test_safety_and_ordered_columns_say_so(pack, base):
    for sheet in pack.sheets:
        for column in sheet.columns:
            if column.kind != "attribute":
                continue
            definition = base.attr_defs[column.name]
            assert column.safety == definition.safety_class
            if definition.safety_class:
                assert "safety" in column.marker()
            if definition.ordered:
                assert "do not sort" in column.marker()


def test_the_derivation_is_stable(base):
    assert schema.columns_for("food", base) == schema.columns_for("food", base)


# ---------------------------------------------------------------------------
# The worked example
# ---------------------------------------------------------------------------


def test_the_example_is_one_suppliers_lines(pack, base):
    """A bundle is one supplier's submission, so an example must be too.

    An example drawn from across a branch would be mostly rejected on arrival,
    which teaches the wrong thing about the format and the right thing about
    nothing.
    """
    for sheet in pack.sheets:
        example = sample_mod.build(sheet, base)
        if not example.rows:
            continue
        suppliers = {
            base.products[base.product_of_variant[v]].supplier
            for v in base.variants
            if base.variants[v].sku
            and any(r["sku"] == base.variants[v].sku for r in example.rows)
        }
        assert suppliers == {example.supplier}, sheet.branch


def test_every_representable_defect_is_shown_or_explained(pack, base):
    """Five of the seven, and the pack says which two it cannot show.

    Shipping five and calling them seven would be the small dishonesty this
    whole system is arranged to avoid. A branch that has no column to
    demonstrate one on records that, and the README prints it.
    """
    shown: set[str] = set()
    for sheet in pack.sheets:
        example = sample_mod.build(sheet, base)
        for row in example.rows:
            note = row.get(sample_mod.NOTE_COLUMN, "")
            for defect in sample_mod.REPRESENTABLE:
                if note.startswith(defect.value):
                    shown.add(defect.value)
        for defect in example.unshown:
            assert defect in {d.value for d in sample_mod.REPRESENTABLE}

    assert shown == {d.value for d in sample_mod.REPRESENTABLE}
    assert set(sample_mod.NOT_REPRESENTABLE) | set(sample_mod.REPRESENTABLE) == set(Defect)


def test_a_broken_example_row_is_broken_in_exactly_one_way(pack, base):
    """Two defects on one row demonstrates neither."""
    for sheet in pack.sheets:
        example = sample_mod.build(sheet, base)
        for row in example.rows:
            note = row.get(sample_mod.NOTE_COLUMN, "")
            if not note:
                continue
            named = [d for d in sample_mod.REPRESENTABLE
                     if note.startswith(d.value)]
            assert len(named) == 1, note


def test_the_example_never_blanks_a_safety_field_to_make_a_point(pack, base):
    """The one column a worked example must not teach by removing.

    A supplier who copies the example row is a supplier who has copied a row
    with no allergen declaration on it.
    """
    for sheet in pack.sheets:
        safety = [c.name for c in sheet.columns if c.safety]
        example = sample_mod.build(sheet, base)
        for row in example.rows:
            note = row.get(sample_mod.NOTE_COLUMN, "")
            if note.startswith(Defect.MISSING_MANDATORY.value):
                assert not any(f"{name} is blank" in note for name in safety)


# ---------------------------------------------------------------------------
# The formats
# ---------------------------------------------------------------------------


def test_csv_round_trips_through_its_own_reader(pack, base):
    sheet = pack.sheet("food")
    example = sample_mod.build(sheet, base)
    text = csv_txt.write_csv(sheet, example)
    rows = csv_txt.read_rows(text)
    assert len(rows) == len(example.rows)
    assert rows[0]["sku"] == example.rows[0]["sku"]
    # The list separator survives, and a comma inside a value does not split it.
    assert schema.LIST_SEPARATOR.strip() in rows[0]["food.ingredients"]


def test_the_two_annotation_rows_are_not_read_as_products(pack, base):
    sheet = pack.sheet("food")
    text = csv_txt.write_csv(sheet, sample_mod.build(sheet, base))
    rows = csv_txt.read_rows(text)
    assert all(r["sku"].lower() != "sku" for r in rows)


def test_a_row_with_no_sku_is_kept_for_the_reader_to_refuse(pack, base):
    """Not dropped here. Silently losing it would report a bundle of
    thirty-nine as complete when the supplier sent forty."""
    sheet = pack.sheet("food")
    text = csv_txt.write_csv(sheet)
    names = [c.name for c in sheet.columns]
    blank = ",".join("" for _ in names)
    rows = csv_txt.read_rows(text + blank + "\r\n" + blank.replace(",", "x,", 1))
    assert any(not r["sku"] for r in rows)


def test_the_pipe_file_carries_the_same_header(pack, base):
    sheet = pack.sheet("electronics")
    text = csv_txt.write_txt(sheet)
    lines = text.splitlines()
    assert len(lines) == 3
    assert [c.strip() for c in lines[0].split("|")] == [c.name for c in sheet.columns]


def test_json_schema_enumerates_the_categories_and_types(pack):
    doc = jsonschema.write(pack)
    for sheet in pack.sheets:
        body = doc["$defs"][sheet.branch]
        assert body["properties"]["category"]["enum"] == list(sheet.leaves)
        assert "sku" in body["required"]
        for column in sheet.columns:
            if column.kind == "attribute":
                assert column.name in body["properties"]
    # The GTIN pattern is the one Excel breaks.
    gtin = doc["$defs"]["food"]["properties"]["identifiers.gtin"]
    assert gtin["type"] == "string"
    assert "pattern" in gtin


def test_the_specification_is_a_readable_docx_naming_every_attribute(pack, base):
    payload = specdoc.write(pack)
    archive = zipfile.ZipFile(io.BytesIO(payload))
    assert archive.testzip() is None
    for part in ("[Content_Types].xml", "_rels/.rels", "word/document.xml",
                 "word/styles.xml"):
        assert part in archive.namelist()
    document = ElementTree.fromstring(archive.read("word/document.xml"))
    text = " ".join(e.text or "" for e in document.iter(f"{W}t"))
    for path, definition in base.attr_defs.items():
        assert path in text, path
        assert definition.label in text, definition.label


def test_the_specification_is_byte_identical_between_builds(pack):
    assert specdoc.write(pack) == specdoc.write(pack)


@pytest.mark.skipif(not workbook.available(), reason="openpyxl is not installed")
def test_the_workbook_keeps_the_gtin_column_text(pack, base):
    """The trap this format exists to close.

    Excel reads 05012345600018 as a number, drops the leading zero, and offers
    to render the rest in scientific notation. The zero is part of the number.
    """
    import openpyxl

    examples = {s.branch: sample_mod.build(s, base) for s in pack.sheets}
    example = examples["food"]
    book = openpyxl.load_workbook(io.BytesIO(workbook.write(pack, examples)))
    sheet = book["Grocery"]
    names = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    column = names.index("identifiers.gtin") + 1

    # The column is text, so nothing in it is reinterpreted as a number.
    assert sheet.cell(5, column).number_format == "@"

    # And every value survives as the string the catalog holds. Asserted
    # against the source rather than against a leading zero: whether this
    # supplier's GTINs happen to start with one is an accident of the seed,
    # and a test that depended on it would pass for the wrong reason and fail
    # when the draw moved.
    for offset, row in enumerate(example.rows, start=4):
        written = sheet.cell(offset, column).value
        assert written == (row.get("identifiers.gtin") or None) or (
            written == "" and not row.get("identifiers.gtin"))
        assert written is None or isinstance(written, str)

    # The trap itself, stated directly: a leading zero is not lost.
    with_zero = dict(example.rows[0])
    with_zero["identifiers.gtin"] = "05012345600018"
    probe = sample_mod.Example(branch="food", supplier=example.supplier,
                               rows=[with_zero], images={})
    reread = openpyxl.load_workbook(
        io.BytesIO(workbook.write(pack, {"food": probe})))["Grocery"]
    assert reread.cell(4, column).value == "05012345600018"


@pytest.mark.skipif(not workbook.available(), reason="openpyxl is not installed")
def test_the_workbook_offers_a_dropdown_where_the_catalog_has_a_vocabulary(pack, base):
    import openpyxl

    book = openpyxl.load_workbook(io.BytesIO(workbook.write(pack, {})))
    assert book["Lists"].sheet_state == "hidden"
    validated = 0
    for sheet in book.worksheets:
        validated += len(sheet.data_validations.dataValidation)
    assert validated > 0, "the catalog has string attributes with settled values"


def test_the_formats_report_what_this_installation_can_actually_build():
    from sc.datapack import formats

    reported = formats()
    assert reported["csv"]["available"] is True
    assert reported["docx"]["available"] is True
    assert reported["xlsx"]["available"] is workbook.available()
    if not workbook.available():
        assert "openpyxl" in reported["xlsx"]["why"]
